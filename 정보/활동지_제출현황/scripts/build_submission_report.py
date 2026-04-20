import json
import re
import sys
import csv
import hashlib
from datetime import datetime, date
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
STUDENTS_JSON = ROOT / "data" / "students_grade1.json"
EXCEL_PATH = Path(r"C:/Users/User/Downloads/2026 정보 활동지.xlsx")
OUTPUT_CSV = ROOT / "활동지_제출현황.csv"

SHEETS = [
    {
        "name": "인공지능시대",
        "sheet_name": "인공지능 시대",
        "combined_id_col": "학번과 이름을 적어주세요",
        "ts_col": "타임스탬프",
    },
    {
        "name": "대형언어모델",
        "sheet_name": "대형언어모델",
        "combined_id_col": "이름과 학번",
        "ts_col": "타임스탬프",
    },
    {
        "name": "자기평가문",
        "sheet_name": "자기평가문활동지",
        "name_col": "이름",
        "classnum_col": "반/번호",
        "ts_col": "제출시각",
    },
    {
        "name": "에이전틱AI",
        "sheet_name": "에이전틱AI활동지",
        "name_col": "이름",
        "classnum_col": "반/번호",
        "ts_col": " ",
    },
    {
        "name": "이미지분류",
        "sheet_name": "이미지분류프로젝트활동지",
        "name_col": "이름",
        "classnum_col": "반/번호",
        "ts_col": "제출시각",
    },
]

HANGUL_RE = re.compile(r"[가-힣]+")
DIGIT_RE = re.compile(r"\d+")


def parse_classnum(s):
    """Parse various class/number formats. Returns (class, number) or (None, None).

    Examples:
        "3반 13번"  -> (3, 13)
        "3-13"      -> (3, 13)
        "1반22"     -> (1, 22)
        "119"       -> (1, 19)
        "10313"     -> (3, 13)  # 5 digits: grade(1) + class(2) + number(2)
        "1315"      -> (3, 15)  # 4 digits: grade(1) + class(1 or 2) + number(2)
        "313"       -> (3, 13)  # 3 digits: class(1) + number(2)
        "10119.0"   -> (1, 19)  # trailing .0 from float
    """
    if s is None:
        return (None, None)
    if isinstance(s, (datetime, date)):
        # date value leaked into class/number column
        return (None, None)
    if isinstance(s, float):
        s = str(int(s)) if s == int(s) else str(s)
    s = str(s).strip()
    if not s:
        return (None, None)

    # Strip trailing ".0" from Excel-stored numerics
    s_clean = re.sub(r"\.0+$", "", s)

    digits_groups = DIGIT_RE.findall(s_clean)
    if not digits_groups:
        return (None, None)

    # Two or more numeric tokens: "3반 13번", "1-22", "2026-03-05" (date form filtered elsewhere)
    if len(digits_groups) >= 2:
        try:
            a, b = int(digits_groups[0]), int(digits_groups[1])
            # If a looks like a full student ID (>= 1000), fall through to single-token parser
            if a >= 100:
                concat = "".join(digits_groups)
            else:
                return (a, b)
        except ValueError:
            return (None, None)
    else:
        concat = digits_groups[0]

    if len(concat) == 5:
        # grade(1) + class(2) + number(2): "10313" -> (3, 13)
        return (int(concat[1:3]), int(concat[3:5]))
    if len(concat) == 4:
        # Ambiguous: grade(1)+class(1)+number(2) = "1315" -> (3,15)
        # OR class(2)+number(2) = "1315" -> (13,15)
        # Prefer grade-prefixed interpretation if first digit is 1 and class 1-9 mapped student exists.
        # Here class range is 1-9 so "1315" most likely (3, 15).
        g, rest = int(concat[0]), concat[1:]
        if g == 1:
            return (int(rest[0]), int(rest[1:3]))
        # Fallback: first 2 digits as class, last 2 as number
        return (int(concat[0:2]), int(concat[2:4]))
    if len(concat) == 3:
        return (int(concat[0]), int(concat[1:3]))
    if len(concat) == 2:
        return (int(concat[0]), int(concat[1]))
    return (None, None)


def parse_combined_id(s):
    """Parse combined id like '10224 최하진' or '10111 박주연'.

    Returns (name, class, number) with any None for unparseable parts.
    """
    if s is None:
        return (None, None, None)
    if isinstance(s, float):
        s = str(int(s)) if s == int(s) else str(s)
    s = str(s).strip()
    if not s:
        return (None, None, None)

    name_match = HANGUL_RE.findall(s)
    name = "".join(name_match) if name_match else None

    s_clean = re.sub(r"\.0+$", "", s)
    digits_groups = DIGIT_RE.findall(s_clean)
    cls, num = (None, None)
    if digits_groups:
        digits_str = "".join(digits_groups)
        if len(digits_str) == 5:
            cls, num = int(digits_str[1:3]), int(digits_str[3:5])
        elif len(digits_str) == 4:
            g, rest = int(digits_str[0]), digits_str[1:]
            if g == 1:
                cls, num = int(rest[0]), int(rest[1:3])
            else:
                cls, num = int(digits_str[0:2]), int(digits_str[2:4])
        elif len(digits_str) == 3:
            cls, num = int(digits_str[0]), int(digits_str[1:3])
        elif len(digits_groups) >= 2:
            cls, num = int(digits_groups[0]), int(digits_groups[1])
    return (name, cls, num)


def parse_name(s):
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None
    matches = HANGUL_RE.findall(s)
    return "".join(matches) if matches else s


def parse_timestamp(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat(timespec="seconds")
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    formats = [
        "%Y. %m. %d. %p %I:%M:%S",
        "%Y. %m. %d. 오전 %I:%M:%S",
        "%Y. %m. %d. 오후 %I:%M:%S",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
    ]
    s_try = s.replace("AM", "오전").replace("PM", "오후")
    for fmt in formats:
        try:
            if "오전" in fmt or "오후" in fmt:
                if "오전" in s_try:
                    parsed = datetime.strptime(s_try, fmt)
                elif "오후" in s_try:
                    parsed = datetime.strptime(s_try, fmt)
                    if parsed.hour != 12:
                        parsed = parsed.replace(hour=parsed.hour + 12)
                else:
                    continue
                return parsed.isoformat(timespec="seconds")
            return datetime.strptime(s, fmt).isoformat(timespec="seconds")
        except ValueError:
            continue
    return s


def normalize_cell(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return re.sub(r"\s+", " ", str(v)).strip()


INFORMATICS_CLASSES = {1, 2, 3, 4}  # 5~9반은 정보 수업 미이수


def load_students():
    with STUDENTS_JSON.open(encoding="utf-8") as f:
        students = json.load(f)
    return [s for s in students if s["class"] in INFORMATICS_CLASSES]


def build_student_index(students):
    index = {}
    by_name = defaultdict(list)
    by_classnum = {}
    for st in students:
        key = (st["name"], st["class"], st["number"])
        index[key] = st
        by_name[st["name"]].append(key)
        by_classnum[(st["class"], st["number"])] = key
    return index, by_name, by_classnum


def name_similarity(a, b):
    """Rough similarity: shared leading/trailing chars."""
    if not a or not b:
        return 0
    if a == b:
        return 1.0
    # Prefix match ratio
    m = min(len(a), len(b))
    prefix = sum(1 for i in range(m) if a[i] == b[i])
    return prefix / max(len(a), len(b))


def resolve_key(name, cls, num, index, by_name, by_classnum):
    """Try exact match, then fallbacks. Returns (key or None, resolution_note)."""
    if name is not None and cls is not None and num is not None:
        k = (name, cls, num)
        if k in index:
            return k, "exact"
    # Class/number valid → unique student at that slot wins
    if cls is not None and num is not None:
        k = by_classnum.get((cls, num))
        if k is not None:
            if name is None or k[0] == name:
                return k, "by_classnum"
            # Name mismatch: allow if names overlap (truncation/typo of 2+ chars)
            if name_similarity(name, k[0]) >= 0.5:
                return k, "by_classnum_fuzzy"
    if name is not None:
        candidates = by_name.get(name, [])
        if len(candidates) == 1:
            return candidates[0], "by_name_unique"
        if len(candidates) > 1 and cls is not None:
            for c in candidates:
                if c[1] == cls:
                    return c, "by_name_class"
    return None, "unmatched"


def process_sheet(ws, sheet_cfg, student_index, by_name, by_classnum):
    headers = [c.value for c in ws[1]]
    header_idx = {h: i for i, h in enumerate(headers) if h is not None}

    ts_col = sheet_cfg["ts_col"]
    ts_idx = header_idx.get(ts_col)

    combined = "combined_id_col" in sheet_cfg
    if combined:
        id_idx = header_idx.get(sheet_cfg["combined_id_col"])
        name_idx = classnum_idx = None
    else:
        id_idx = None
        name_idx = header_idx.get(sheet_cfg["name_col"])
        classnum_idx = header_idx.get(sheet_cfg["classnum_col"])

    answer_indices = []
    for i, h in enumerate(headers):
        if h is None:
            continue
        if combined:
            if h in (ts_col, sheet_cfg["combined_id_col"]):
                continue
        else:
            if h in (ts_col, sheet_cfg["name_col"], sheet_cfg["classnum_col"]):
                continue
        answer_indices.append(i)

    submissions_by_student = defaultdict(list)
    unmatched = []
    hash_to_students = defaultdict(set)
    student_hashes = defaultdict(set)

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue

        if combined:
            raw_id = row[id_idx] if id_idx is not None else None
            name, cls, num = parse_combined_id(raw_id)
            raw_label = raw_id
        else:
            name_raw = row[name_idx] if name_idx is not None else None
            classnum_raw = row[classnum_idx] if classnum_idx is not None else None
            name = parse_name(name_raw)
            cls, num = parse_classnum(classnum_raw)
            raw_label = f"{name_raw} | {classnum_raw}"

        ts_raw = row[ts_idx] if ts_idx is not None else None
        ts = parse_timestamp(ts_raw)

        key, note = resolve_key(name, cls, num, student_index, by_name, by_classnum)
        matched = key is not None

        answer_values = [normalize_cell(row[i]) for i in answer_indices]
        content_hash = hashlib.sha256("\x1f".join(answer_values).encode("utf-8")).hexdigest()
        char_count = sum(len(v) for v in answer_values if not v.startswith(("http://", "https://")))

        if matched:
            submissions_by_student[key].append({"time": ts, "chars": char_count})
            hash_to_students[content_hash].add(key)
            student_hashes[key].add(content_hash)
        else:
            unmatched.append({
                "row": row_num,
                "raw": raw_label,
                "parsed": (name, cls, num),
                "timestamp": ts,
                "note": note,
            })

    duplicate_groups_by_student = defaultdict(list)
    for h, students in hash_to_students.items():
        if len(students) >= 2:
            for st in students:
                others = [s for s in students if s != st]
                for o in others:
                    label = f"1{o[1]:02d}{o[2]:02d} {o[0]}"
                    if label not in duplicate_groups_by_student[st]:
                        duplicate_groups_by_student[st].append(label)

    return {
        "submissions": submissions_by_student,
        "unmatched": unmatched,
        "duplicate_groups": duplicate_groups_by_student,
    }


def main():
    students = load_students()
    student_index, by_name, by_classnum = build_student_index(students)
    print(f"Loaded {len(students)} students", file=sys.stderr)

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    per_sheet_results = {}
    for cfg in SHEETS:
        if cfg["sheet_name"] not in wb.sheetnames:
            print(f"WARN: sheet not found: {cfg['sheet_name']}", file=sys.stderr)
            per_sheet_results[cfg["name"]] = {"submissions": {}, "unmatched": [], "duplicate_groups": {}}
            continue
        ws = wb[cfg["sheet_name"]]
        result = process_sheet(ws, cfg, student_index, by_name, by_classnum)
        per_sheet_results[cfg["name"]] = result
        total_submissions = sum(len(v) for v in result["submissions"].values())
        print(
            f"[{cfg['name']}] matched_submissions={total_submissions} "
            f"unique_students={len(result['submissions'])} unmatched={len(result['unmatched'])}",
            file=sys.stderr,
        )
        for u in result["unmatched"]:
            print(
                f"  UNMATCHED [{cfg['name']}] row={u['row']} raw={u['raw']!r} parsed={u['parsed']}",
                file=sys.stderr,
            )

    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        sheet_names = [cfg["name"] for cfg in SHEETS]
        writer.writerow(
            ["grade", "class", "number", "name"] + sheet_names + ["duplicate_groups"]
        )

        for st in sorted(students, key=lambda s: (s["class"], s["number"])):
            key = (st["name"], st["class"], st["number"])
            row = [1, st["class"], st["number"], st["name"]]

            duplicate_groups = {}
            for cfg in SHEETS:
                result = per_sheet_results[cfg["name"]]
                subs = result["submissions"].get(key, [])
                cell = json.dumps(
                    {
                        "count": len(subs),
                        "times": [s["time"] for s in subs],
                        "chars": [s["chars"] for s in subs],
                    },
                    ensure_ascii=False,
                )
                row.append(cell)
                dup = result["duplicate_groups"].get(key, [])
                if dup:
                    duplicate_groups[cfg["name"]] = dup

            row.append(json.dumps(duplicate_groups, ensure_ascii=False))
            writer.writerow(row)

    print(f"\nWritten: {OUTPUT_CSV}", file=sys.stderr)


if __name__ == "__main__":
    main()
